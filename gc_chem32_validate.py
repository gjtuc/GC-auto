# -*- coding: utf-8 -*-
"""
gc_chem32_validate.py — GC3 Chem32 실데이터 로컬 검증 (GC8860 / Cursor)

GC3 PC에서 cmd가 바로 닫혀도 Desktop\\KCH\\gc3_validate_log.txt 에 전체 로그 저장.

사용:
  python gc_chem32_validate.py --data-path testdata\\gc3_e2e\\Chem32\\1\\Data
  python gc_chem32_validate.py --data-path ... --sample-folder "20260620 DRE..."
  python gc_chem32_validate.py --data-path ... --audit --compare-xlsx Downloads\\....xlsx
  python gc_chem32_validate.py --data-path ... --run-pipeline --no-email
"""

from __future__ import annotations

import argparse
import os
import sys
import traceback
from contextlib import redirect_stdout
from datetime import datetime
from io import StringIO
from typing import Dict, List, Optional, Tuple

from gc_console import setup_console_encoding

setup_console_encoding()

from gc_chem32 import (
    build_merged_injection_cycles,
    collect_reported_injections,
    cycles_match,
    default_sample_name_from_folder,
    describe_cycle_mismatch,
    detect_analysis_gaps,
    find_active_sample_folder,
    find_chem32_injection_folders,
    find_report_txt,
    find_sample_folders,
    find_sequence_folders,
    format_duration_korean,
    get_latest_sequence_datetime,
    parse_injection_reports,
    parse_report_injection_datetime,
    _resolve_reference_index,
)
from gc_config import AREA_MATCH_TOLERANCE, DEFAULT_GC3_DATA, EXCEL_OUTPUT_DIR, RT_TOLERANCE
from gc_profiles import resolve_data_path


class _Tee:
    def __init__(self, *streams):
        self._streams = streams

    def write(self, data):
        for stream in self._streams:
            stream.write(data)

    def flush(self):
        for stream in self._streams:
            stream.flush()


def _log_path() -> str:
    out = os.path.join(os.path.expanduser("~"), "Desktop", "KCH")
    os.makedirs(out, exist_ok=True)
    return os.path.join(out, "gc3_validate_log.txt")


def count_xlsx_cycles(xlsx_path: str) -> Dict[str, int]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    counts: Dict[str, int] = {}
    for sheet_name in wb.sheetnames:
        worksheet = wb[sheet_name]
        cycles = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row and row[0] == "#":
                cycles += 1
        counts[sheet_name] = cycles
    return counts


def _print_sample_tree(sample_folder: str) -> None:
    print(f"\n=== 시료 폴더 ===\n{sample_folder}\n")
    sequences = find_sequence_folders(sample_folder)
    print(f"REACTION 시퀀스 {len(sequences)}개:")
    for seq in sequences:
        print(f"  - {os.path.basename(seq)}")
        for inj in find_chem32_injection_folders(seq):
            label = os.path.basename(inj)
            report = find_report_txt(inj)
            reports = parse_injection_reports(inj)
            fid_n = len(reports.get("FID", []))
            tcd_n = len(reports.get("TCD", []))
            rep = os.path.basename(report) if report else "(없음)"
            print(f"      {label}  Report={rep}  FID={fid_n}  TCD={tcd_n}")


def _audit_sliding_chain(sample_folder: str, detector_key: str = "TCD") -> None:
    """시퀀스별 sliding chain — 주입마다 직전 대비 통과/실패."""
    from itertools import groupby

    injections = collect_reported_injections(sample_folder)
    print(f"\n=== AUDIT {detector_key} sliding chain (직전 주입 대비) ===")
    print(
        f"허용: RT ±{RT_TOLERANCE} min, Area ±{AREA_MATCH_TOLERANCE * 100:.0f}% (인접 주입)"
    )

    total_ok = 0
    total_skip = 0
    for seq_path, group in groupby(injections, key=lambda item: item[1]):
        seq_name = os.path.basename(seq_path)
        chunk = list(group)
        labels = [os.path.basename(path) for path, _ in chunk]
        peaks_list = [
            parse_injection_reports(path).get(detector_key, []) for path, _ in chunk
        ]
        if not any(peaks_list):
            continue

        ref_index = _resolve_reference_index(peaks_list, labels, detector_key)
        print(f"\n--- {seq_name} (주입 {len(chunk)}개, 기준={labels[ref_index]}) ---")

        chain_ref: Optional[List[dict]] = None
        for index, (label, peaks) in enumerate(zip(labels, peaks_list)):
            if not peaks:
                print(f"  [없음] {label}: {detector_key} 피크 0")
                total_skip += 1
                continue
            if index < ref_index:
                print(f"  [startup] {label}: {len(peaks)}피크")
                total_skip += 1
                continue
            if index == ref_index:
                chain_ref = peaks
                print(f"  [기준] {label}: {len(peaks)}피크")
                total_ok += 1
                continue
            assert chain_ref is not None
            if len(peaks) != len(chain_ref):
                print(
                    f"  [제외] {label}: 피크 수 {len(peaks)} "
                    f"(직전 {len(chain_ref)})"
                )
                total_skip += 1
                continue
            if cycles_match(chain_ref, peaks):
                print(f"  [OK] {label}: {len(peaks)}피크")
                chain_ref = peaks
                total_ok += 1
            else:
                reason = describe_cycle_mismatch(chain_ref, peaks)
                print(f"  [제외] {label}: {reason}")
                total_skip += 1

    print(f"\n[AUDIT 합계] {detector_key} 통과 {total_ok} / 제외 {total_skip}")


def _audit_gap_injections(sample_folder: str, gaps, injections) -> None:
    """분석 중단 구간 전후 Injection Date — Chem32 원본 대조용."""
    if not gaps:
        return
    print("\n=== 갭 주입 시각 감사 (Report Injection Date) ===")
    for gap_index, gap in enumerate(gaps, start=1):
        print(
            f"\n갭 {gap_index}: #{gap.after_injection_index + 1} → #{gap.before_injection_index + 1}  "
            f"({gap.after_sequence} → {gap.before_sequence})"
        )
        lo = max(0, gap.after_injection_index - 2)
        hi = min(len(injections), gap.before_injection_index + 3)
        for pos in range(lo, hi):
            inj_path, seq_path = injections[pos]
            report = find_report_txt(inj_path)
            dt = parse_report_injection_datetime(report) if report else None
            mark = ""
            if pos == gap.after_injection_index:
                mark = "  ← 갭 직전"
            elif pos == gap.before_injection_index:
                mark = "  ← 갭 직후"
            when = dt.strftime("%Y-%m-%d %H:%M:%S") if dt else "Injection Date 없음"
            print(
                f"  #{pos + 1:3d} {os.path.basename(inj_path):12s}  {when}  "
                f"({os.path.basename(seq_path)[:48]}){mark}"
            )


def _print_compare_xlsx(
    xlsx_path: str,
    fid_cycles: int,
    tcd_cycles: int,
    raw_injections: int,
) -> None:
    if not os.path.isfile(xlsx_path):
        print(f"\n[오류] xlsx 없음: {xlsx_path}")
        return
    counts = count_xlsx_cycles(xlsx_path)
    print(f"\n=== xlsx 대조: {xlsx_path} ===")
    for sheet, cycles in counts.items():
        print(f"  시트 {sheet}: {cycles} 사이클")
    fid_x = counts.get("FID", counts.get("Sheet1", 0))
    tcd_x = counts.get("TCD", 0)
    print(f"\n  Report 주입(원본)     : {raw_injections}")
    print(f"  pipeline FID/TCD      : {fid_cycles} / {tcd_cycles}")
    print(f"  xlsx FID/TCD          : {fid_x} / {tcd_x}")
    if fid_x and fid_cycles != fid_x:
        print(f"  [불일치] FID pipeline {fid_cycles} vs xlsx {fid_x} (차이 {fid_cycles - fid_x:+d})")
    elif fid_x:
        print("  [OK] FID 사이클 수 일치")
    if tcd_x and tcd_cycles != tcd_x:
        print(f"  [불일치] TCD pipeline {tcd_cycles} vs xlsx {tcd_x} (차이 {tcd_cycles - tcd_x:+d})")
    elif tcd_x:
        print("  [OK] TCD 사이클 수 일치")
    report_gap = raw_injections - fid_cycles
    if report_gap > 2:
        print(
            f"\n  [힌트] Report {raw_injections}개 중 pipeline {fid_cycles}사이클 "
            f"(차이 {report_gap}) — startup·미완료 주입 제외. "
            "xlsx 가 구버전이면 [불일치] 행의 차이와 혼동하지 마세요."
        )
    if fid_x and abs(fid_cycles - fid_x) >= 5:
        print(
            f"\n  [힌트] xlsx FID {fid_x} vs pipeline {fid_cycles} — "
            "구 sliding 코드로 만든 엑셀이면 차이가 큽니다. "
            "최신 코드로 pipeline 재실행 후 Desktop\\KCH xlsx 와 비교하세요."
        )


def run_validate(args: argparse.Namespace) -> int:
    data_path = os.path.normpath(args.data_path or resolve_data_path() or DEFAULT_GC3_DATA)
    print(f"[Data] {data_path}")
    if not os.path.isdir(data_path):
        print(f"[오류] Data 경로 없음: {data_path}")
        return 1

    samples = find_sample_folders(data_path)
    print(f"[안내] 시료 후보 {len(samples)}개 (최신 시퀀스 시각순):")
    for index, path in enumerate(samples[:10]):
        latest = get_latest_sequence_datetime(path)
        when = latest.strftime("%Y-%m-%d %H:%M:%S") if latest else "시퀀스 시각 없음"
        print(f"  {index + 1}. {os.path.basename(path)}  최신시퀀스={when}")

    sample_folder = None
    if args.sample_folder:
        candidate = args.sample_folder
        if not os.path.isabs(candidate):
            candidate = os.path.join(data_path, candidate)
        sample_folder = os.path.normpath(candidate)
        if not os.path.isdir(sample_folder):
            print(f"[오류] 시료 폴더 없음: {sample_folder}")
            return 1
    else:
        sample_folder = find_active_sample_folder(data_path)

    if not sample_folder:
        print("[오류] 시료 폴더를 찾지 못함")
        return 1

    _print_sample_tree(sample_folder)
    raw_count = len(collect_reported_injections(sample_folder))
    print(f"\n[안내] Report 있는 주입 총 {raw_count}개")

    if args.audit:
        _audit_sliding_chain(sample_folder, "TCD")
        _audit_sliding_chain(sample_folder, "FID")

    print("\n=== 병합 결과 (pipeline 과 동일) ===")
    fid_cycles, tcd_cycles, matched, skipped, matched_paths = build_merged_injection_cycles(sample_folder)
    gaps, gap_interval = detect_analysis_gaps(sample_folder)
    gap_injections = collect_reported_injections(sample_folder)
    if gap_interval:
        print(f"\n사이클 간격 추정: {format_duration_korean(gap_interval)}")
    if gaps:
        total_missing = sum(g.missing_cycles for g in gaps)
        print(f"분석 중단 구간 {len(gaps)}곳 — 추정 미수집 약 {total_missing}사이클 (floor, 잔여 버림)")
        from gc_chem32 import analysis_gaps_email_lines

        for line in analysis_gaps_email_lines(gaps, gap_interval, gap_injections):
            print(line)
        _audit_gap_injections(sample_folder, gaps, gap_injections)
    default_name = default_sample_name_from_folder(sample_folder)
    print(f"시료명(자동): {default_name!r}")
    print(
        f"FID 사이클 {len(fid_cycles)} / TCD 사이클 {len(tcd_cycles)} / "
        f"매칭 주입 {len(matched)} / 건너뜀 {skipped}"
    )
    if len(fid_cycles) != len(tcd_cycles):
        print(
            f"\n[오류] FID·TCD 사이클 수 불일치 — "
            f"완료 주입은 1주입당 FID+TCD 1쌍이어야 함"
        )
        return 1

    if args.compare_xlsx:
        xlsx_path = args.compare_xlsx
        if not os.path.isabs(xlsx_path):
            xlsx_path = os.path.join(os.path.expanduser("~"), xlsx_path)
        _print_compare_xlsx(
            os.path.normpath(xlsx_path),
            len(fid_cycles),
            len(tcd_cycles),
            raw_count,
        )

    if skipped >= 5:
        print(
            "\n[힌트] 건너뜀 다수 — GC3 코드가 구버전(sliding)일 수 있습니다.\n"
            "  · 최신: startup·미완료만 제외 (보통 1~2개)\n"
            "  · gc3_make_deploy_zip.bat 후 GC3 PC에 zip 재배포\n"
            "  · --audit 는 구 sliding 진단용(현재 pipeline 과 무관)"
        )

    if args.run_pipeline:
        print("\n=== pipeline 실행 (--no-email) ===")
        from gc_automation import config_from_args, apply_env_overrides
        from gc_pipeline import run_processing
        from gc_config import AppConfig
        from dataclasses import replace

        script_dir = os.path.dirname(os.path.abspath(__file__))
        ns = argparse.Namespace(
            data_path=data_path,
            chemstation_mode="chem32",
            sample_name=args.sample_name,
            sequence_folder=sample_folder,
            sequence_date=None,
            detector="TCD",
            no_email=True,
            no_wifi_check=True,
            force=True,
            watch=False,
            required_ssid=None,
            send_state_file=None,
            watch_status_json=None,
            watch_status_txt=None,
        )
        config = apply_env_overrides(
            config_from_args(ns, script_dir),
            script_dir,
            chemstation_mode_cli="chem32",
        )
        config = replace(config, sequence_folder=sample_folder)
        result = run_processing(config, script_dir)
        print(f"\n결과: ok={result.ok}  reason={result.fail_reason or '-'}")
        if result.output_path:
            print(f"엑셀: {result.output_path}")
            if args.compare_xlsx:
                _print_compare_xlsx(
                    os.path.normpath(
                        args.compare_xlsx
                        if os.path.isabs(args.compare_xlsx)
                        else os.path.join(os.path.expanduser("~"), args.compare_xlsx)
                    ),
                    len(fid_cycles),
                    len(tcd_cycles),
                    raw_count,
                )
        return 0 if result.ok else 1

    if not fid_cycles and not tcd_cycles:
        return 1
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="GC3 Chem32 실데이터 검증")
    parser.add_argument("--data-path", default=None, help=f"기본: {DEFAULT_GC3_DATA}")
    parser.add_argument("--sample-folder", default=None, help="시료 폴더 이름 또는 전체 경로")
    parser.add_argument("--sample-name", default=None, help="pipeline 테스트 시 시료명")
    parser.add_argument(
        "--audit",
        action="store_true",
        help="주입별 sliding chain 통과/실패 (피크#·Area% 사유)",
    )
    parser.add_argument(
        "--compare-xlsx",
        default=None,
        help="생성 xlsx 사이클 수와 pipeline 결과 대조",
    )
    parser.add_argument("--run-pipeline", action="store_true", help="엑셀까지 생성 (--no-email)")
    args = parser.parse_args()

    log_file = _log_path()
    buffer = StringIO()
    code = 1
    try:
        with open(log_file, "w", encoding="utf-8") as log_fp:
            tee = _Tee(sys.stdout, buffer, log_fp)
            old_stdout = sys.stdout
            sys.stdout = tee
            try:
                print(f"=== gc_chem32_validate {datetime.now().isoformat(timespec='seconds')} ===")
                code = run_validate(args)
            finally:
                sys.stdout = old_stdout
        print(f"\n[저장] 전체 로그: {log_file}")
    except Exception:
        with open(log_file, "a", encoding="utf-8") as log_fp:
            log_fp.write("\n[예외]\n")
            log_fp.write(traceback.format_exc())
        print(traceback.format_exc())
        print(f"\n[저장] {log_file}")
        code = 1
    sys.exit(code)


if __name__ == "__main__":
    main()
